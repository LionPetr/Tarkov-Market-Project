#test

from urllib.request import urlopen
import re
from openpyxl import load_workbook

workbook = load_workbook(filename="excel work.xlsx")
workbook.sheetnames

sheet = workbook.active
url = "https://tarkov-market.com/tag/tactical_rigs"
page = urlopen(url)
html_source = page.read()
html = html_source.decode("utf-8")
displacement = [7, 66, "\""]


i = 0
database = [] #url database
urlArray = [] 
#find all the urls
while i != -1: 
    urlplace = html.find("href=\"/item", i + 10)
    database.append(urlplace)
    i = urlplace

#print just the urls 
for i in database[:-1]:
    fullurl = html[i + displacement[0]: i + displacement[1]]
    x = fullurl.index(displacement[2])
    itemurl = fullurl[: x]
    if itemurl not in urlArray:
        urlArray.append(itemurl)
    else:
        continue
for i in urlArray:
    itemName = i[5:].replace("_", " ")
    sheet["A" + str(urlArray.index(i) + 1)] = i 
    sheet["B" + str(urlArray.index(i) + 1)] = itemName
workbook.save(filename="excel work edited.xlsx")
